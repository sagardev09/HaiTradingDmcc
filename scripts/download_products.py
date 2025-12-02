#!/usr/bin/env python3
"""
Script to process data.json, download product images, and create an Excel sheet
with web_item_name, item_group, description, and both images embedded.
"""

import json
import os
import requests
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from bs4 import BeautifulSoup
import re
from io import BytesIO
from PIL import Image as PILImage

# Configuration
BASE_URL = "https://erp.jogaram.com"
DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "data.json")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "downloaded_images")
OUTPUT_EXCEL = os.path.join(os.path.dirname(__file__), "..", "products_export.xlsx")

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)


def clean_html(html_string):
    """Remove HTML tags and clean up the description text."""
    if not html_string:
        return ""
    soup = BeautifulSoup(html_string, "html.parser")
    text = soup.get_text(separator="\n")
    # Clean up extra whitespace and newlines
    text = re.sub(r'\n\s*\n', '\n', text)
    text = text.strip()
    return text


def download_image(image_path, filename):
    """Download an image from the given path and save it locally (high quality)."""
    if not image_path:
        return None

    # Construct full URL
    full_url = f"{BASE_URL}{image_path}"

    # Encode special characters in the URL path
    # Split URL to encode only the path part
    url_parts = full_url.split("/files/")
    if len(url_parts) == 2:
        encoded_path = quote(url_parts[1], safe='')
        full_url = f"{url_parts[0]}/files/{encoded_path}"

    try:
        print(f"Downloading: {full_url}")
        response = requests.get(full_url, timeout=30)
        response.raise_for_status()

        # Save the image in original quality
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(response.content)

        return filepath
    except requests.exceptions.RequestException as e:
        print(f"Error downloading {full_url}: {e}")
        return None


def main():
    # Load data from JSON file
    print(f"Loading data from {DATA_FILE}...")
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Handle both formats: {"message": [...]} or direct array [...]
    if isinstance(data, dict):
        products = data.get("message", [])
    elif isinstance(data, list):
        products = data
    else:
        products = []
    print(f"Found {len(products)} products")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define headers
    headers = ["S.No", "Web Item Name", "Product Group", "Item Group", "Description", "Main Image", "Thumbnail"]

    # Style for headers
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths (larger for high quality images)
    ws.column_dimensions['A'].width = 8   # S.No
    ws.column_dimensions['B'].width = 40  # Web Item Name
    ws.column_dimensions['C'].width = 20  # Product Group
    ws.column_dimensions['D'].width = 25  # Item Group
    ws.column_dimensions['E'].width = 60  # Description
    ws.column_dimensions['F'].width = 35  # Main Image
    ws.column_dimensions['G'].width = 35  # Thumbnail

    # Process each product
    for idx, product in enumerate(products, 1):
        row = idx + 1  # Account for header row

        print(f"\nProcessing {idx}/{len(products)}: {product.get('web_item_name', 'Unknown')}")

        # Extract data
        web_item_name = product.get('web_item_name', '')
        item_group = product.get('item_group', '')
        description = clean_html(product.get('description', ''))
        website_image = product.get('website_image', '')
        thumbnail = product.get('thumbnail', '')

        # Extract product group from route (first part of the route path)
        route = product.get('route', '')
        product_group = route.split('/')[0].replace('-', ' ').title() if route else ''

        # Write text data
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=web_item_name).border = thin_border
        ws.cell(row=row, column=3, value=product_group).border = thin_border
        ws.cell(row=row, column=4, value=item_group).border = thin_border
        desc_cell = ws.cell(row=row, column=5, value=description)
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Set row height to accommodate high quality images
        ws.row_dimensions[row].height = 180

        # Download and embed main image
        if website_image:
            # Use item_name for filename, get extension from original file
            item_name_safe = re.sub(r'[^a-zA-Z0-9\s\-]', '', product.get('item_name', f'product_{idx}'))
            item_name_safe = re.sub(r'\s+', '_', item_name_safe.strip())[:80]  # Limit length
            ext = os.path.splitext(website_image)[1] or '.jpg'
            safe_filename = f"{item_name_safe}_main{ext}"
            main_img_path = download_image(website_image, safe_filename)
            if main_img_path and os.path.exists(main_img_path):
                try:
                    img = XLImage(main_img_path)
                    img.width = 220
                    img.height = 220
                    ws.add_image(img, f'F{row}')
                except Exception as e:
                    print(f"Error embedding main image: {e}")
                    ws.cell(row=row, column=6, value=f"{BASE_URL}{website_image}").border = thin_border
            else:
                ws.cell(row=row, column=6, value=f"{BASE_URL}{website_image}").border = thin_border

        # Download and embed thumbnail
        if thumbnail:
            # Use item_name for filename, get extension from original file
            item_name_safe = re.sub(r'[^a-zA-Z0-9\s\-]', '', product.get('item_name', f'product_{idx}'))
            item_name_safe = re.sub(r'\s+', '_', item_name_safe.strip())[:80]  # Limit length
            ext = os.path.splitext(thumbnail)[1] or '.jpg'
            safe_filename = f"{item_name_safe}_thumb{ext}"
            thumb_img_path = download_image(thumbnail, safe_filename)
            if thumb_img_path and os.path.exists(thumb_img_path):
                try:
                    img = XLImage(thumb_img_path)
                    img.width = 220
                    img.height = 220
                    ws.add_image(img, f'G{row}')
                except Exception as e:
                    print(f"Error embedding thumbnail: {e}")
                    ws.cell(row=row, column=7, value=f"{BASE_URL}{thumbnail}").border = thin_border
            else:
                ws.cell(row=row, column=7, value=f"{BASE_URL}{thumbnail}").border = thin_border

    # Save the workbook
    print(f"\nSaving Excel file to {OUTPUT_EXCEL}...")
    wb.save(OUTPUT_EXCEL)
    print(f"Done! Excel file saved to: {OUTPUT_EXCEL}")
    print(f"Downloaded images saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
